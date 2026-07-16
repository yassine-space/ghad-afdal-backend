from urllib import request

from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from datetime import timedelta,date
from django.conf import settings             # missing — needed by certificate PDF generation
from django.db.models import Sum, Q          # was: from django.db.models import Sum
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.core.exceptions import ValidationError
from io import BytesIO
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
import barcode
from barcode.writer import ImageWriter
from .serializers import MachineSerializer, MachineAssignmentSerializer
import calendar
from openpyxl import Workbook
from django.http import HttpResponse


from .models import (
    FinancialCategory, 
    Donation,
    ExpenseTransaction,
    FinancialAuditLog,
    DonationHistory,
    Donor,
    Patient,
    Person,
    Department,
    Activity,
    Member,
    Drug,
    DrugStock,
    DrugDonation,
    DonationItem,
    DrugDistribution,
    DistributionItem,
    get_user_activities,
    Machine,
    MachineAssignment
)

from .serializers import (
    DonationHistorySerializer,
    DonorSerializer,
    PatientSerializer,
    UserSerializer,
    PersonSerializer,
    DepartmentSerializer,
    ActivitySerializer,
    MemberSerializer,
    DrugSerializer,
    DrugStockSerializer,
    DrugDonationSerializer,
    DrugDistributionSerializer,
    MachineSerializer,
    MachineAssignmentSerializer,
    FinancialCategorySerializer, 
    DonationSerializer,
    ExpenseTransactionSerializer,
    FinancialAuditLogSerializer,
    )

from .permissions import (
    HasActivityAccess,
    IsAdminOnly,
    IsAuthenticatedAnyActivity,
)


# ─────────────────────────────────────────────
# AUTH VIEWS  
# ─────────────────────────────────────────────

class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    """
    Extends the default JWT login response to include user info
    and their activity accesses — so the frontend knows immediately
    what the user can see after login.
    """
    def validate(self, attrs):
        data  = super().validate(attrs)
        user  = self.user
        data['user'] = UserSerializer(user).data
        return data


class LoginView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class MeView(APIView):
    """
    GET   /api/auth/me/    → current user profile + linked person details
    PATCH /api/auth/me/    → change own password only (username is admin-only)
    """
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'detail': 'Not authenticated.'}, status=status.HTTP_401_UNAUTHORIZED)
        user = (
            User.objects
            .select_related('profile__person')
            .prefetch_related('activity_accesses__activity__department')
            .get(pk=request.user.pk)
        )
        data = UserSerializer(user).data
        try:
            person_obj = user.profile.person if user.profile else None
        except Exception:
            person_obj = None
        if person_obj:
            data['person_detail'] = PersonSerializer(person_obj).data
        else:
            data['person_detail'] = None
        return Response(data)

    def patch(self, request):
        if not request.user.is_authenticated:
            return Response({'detail': 'Not authenticated.'}, status=status.HTTP_401_UNAUTHORIZED)
        from .serializers import MeUpdateSerializer
        serializer = MeUpdateSerializer(
            request.user,
            data=request.data,
            partial=True,
            context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response({'detail': 'تم تغيير كلمة المرور بنجاح.'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
# ─────────────────────────────────────────────
# ADMIN-MANAGED VIEWSETS
# Reads are open to any authenticated user (sidebar/members page need them).
# Writes (POST/PUT/PATCH/DELETE) are admin only.
# ─────────────────────────────────────────────

class DepartmentViewSet(viewsets.ModelViewSet):
    """
    GET             /api/departments/  → any authenticated user
    POST/PUT/DELETE /api/departments/  → admin only
    """
    queryset           = Department.objects.all()
    serializer_class   = DepartmentSerializer

    def get_permissions(self):
        if self.request.method in ('GET', 'HEAD', 'OPTIONS'):
            return [IsAuthenticatedAnyActivity()]
        return [IsAdminOnly()]


class ActivityViewSet(viewsets.ModelViewSet):
    """
    GET             /api/activities/  → any authenticated user
    POST/PUT/DELETE /api/activities/  → admin only
    """
    queryset           = Activity.objects.select_related('department').all()
    serializer_class   = ActivitySerializer

    def get_permissions(self):
        if self.request.method in ('GET', 'HEAD', 'OPTIONS'):
            return [IsAuthenticatedAnyActivity()]
        return [IsAdminOnly()]


class MemberViewSet(viewsets.ModelViewSet):
    """
    GET             /api/members/  → any authenticated user
    POST/PUT/DELETE /api/members/  → admin only
    """
    queryset           = Member.objects.select_related('person', 'department').all()
    serializer_class   = MemberSerializer

    def get_permissions(self):
        if self.request.method in ('GET', 'HEAD', 'OPTIONS'):
            return [IsAuthenticatedAnyActivity()]
        return [IsAdminOnly()]


# ─────────────────────────────────────────────
# SHARED RESOURCE — PERSONS
# Any authenticated user can read persons.
# Only admin can create/edit/delete.
# ─────────────────────────────────────────────

class PersonViewSet(viewsets.ModelViewSet):
    """
    GET/POST/PUT/DELETE /api/persons/ → any authenticated user
    (Regular users need full CRUD on persons via PeoplePage)
    """
    queryset           = Person.objects.all()
    serializer_class   = PersonSerializer
    permission_classes = [IsAuthenticated]


# ─────────────────────────────────────────────
# PHARMACY ACTIVITY VIEWSETS
# These require the user to be assigned to the
# 'Pharmacy' activity inside a department.
#
# HOW TO CONFIGURE:
#   required_activity   = exact name of the Activity in your DB
#   required_department = exact name of the Department in your DB
#
# Change these strings to match what you create in the admin panel.
# ─────────────────────────────────────────────

class DrugViewSet(viewsets.ModelViewSet):
    """
    Required: an activity whose NAME contains one of activity_keywords
    (matched case-insensitively, substring match — so it works no matter
    exactly how you named the activity/department in the admin panel).
    GET           → viewer
    POST/PUT/PATCH → editor
    DELETE        → manager
    """
    queryset            = Drug.objects.all()
    serializer_class    = DrugSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = ['pharma', 'صيدل', 'drug', 'دواء', 'medicine']

    @action(detail=False, methods=['get'], url_path='expiring_soon')
    def expiring_soon(self, request):
        """
        GET /api/drugs/expiring_soon/?days=30
        Returns stock entries expiring within N days.
        Requires viewer access.
        """
        days    = int(request.query_params.get('days', 30))
        cutoff  = timezone.now().date() + timedelta(days=days)
        stocks  = DrugStock.objects.filter(
            expiration_date__lte=cutoff,
            quantity_available__gt=0,
        ).select_related('drug')
        return Response(DrugStockSerializer(stocks, many=True).data)

    @action(detail=True, methods=['get'], url_path='stock')
    def stock(self, request, pk=None):
        """
        GET /api/drugs/{id}/stock/
        Returns all stock batches for a single drug plus lifetime totals.
        Requires viewer access.
        """
        drug = self.get_object()

        # Current available batches (non-zero only)
        stocks = DrugStock.objects.filter(drug=drug, quantity_available__gt=0)
        available = stocks.aggregate(total=Sum('quantity_available'))['total'] or 0

        # Lifetime total received (all donation items for this drug)
        total_received = (
            DonationItem.objects
            .filter(drug=drug)
            .aggregate(total=Sum('quantity'))['total'] or 0
        )

        # Lifetime total distributed (validated distributions only)
        total_distributed = (
            DistributionItem.objects
            .filter(stock__drug=drug, distribution__is_validated=True)
            .aggregate(total=Sum('quantity'))['total'] or 0
        )

        return Response({
            'drug_id':           drug.id,
            'available':         available,
            'total_received':    total_received,
            'total_distributed': total_distributed,
            'batches':           DrugStockSerializer(stocks, many=True).data,
        })


class DrugStockViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only — stock is managed automatically via signals and validate().
    GET /api/stocks/          → all non-zero stock entries
    GET /api/stocks/?drug=5   → filter by drug id
    Requires viewer access on a pharmacy-keyword activity.
    """
    serializer_class    = DrugStockSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = ['pharma', 'صيدل', 'drug', 'دواء', 'medicine']

    def get_queryset(self):
        qs = DrugStock.objects.select_related('drug').filter(quantity_available__gt=0)
        drug_id = self.request.query_params.get('drug')
        if drug_id:
            qs = qs.filter(drug_id=drug_id)
        return qs


class DrugDonationViewSet(viewsets.ModelViewSet):
    """
    GET           → viewer on pharmacy-keyword activity
    POST/PUT/PATCH → editor on pharmacy-keyword activity
    DELETE        → manager on pharmacy-keyword activity
    """
    queryset            = DrugDonation.objects.prefetch_related('items__drug').all()
    serializer_class    = DrugDonationSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = ['pharma', 'صيدل', 'drug', 'دواء', 'medicine']

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel_donation(self, request, pk=None):
        if not request.user.is_superuser:
            from .models import get_user_activities
            LEVEL_RANK = {'viewer': 1, 'editor': 2, 'manager': 3}
            accesses = get_user_activities(request.user)
            has_manager = any(
                any(kw.lower() in (a.activity.name or '').lower() for kw in self.activity_keywords)
                and LEVEL_RANK.get(a.access_level, 0) >= LEVEL_RANK['manager']
                for a in accesses
            )
            if not has_manager:
                return Response(
                    {'detail': 'You need manager access to cancel a donation.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        donation = self.get_object()
        reason = request.data.get('reason', '')
        try:
            donation.cancel(user=request.user, reason=reason)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(DrugDonationSerializer(donation).data)

class DrugDistributionViewSet(viewsets.ModelViewSet):
    """
    GET            → viewer on pharmacy-keyword activity
    POST/PUT/PATCH  → editor on pharmacy-keyword activity
    DELETE         → manager on pharmacy-keyword activity
    validate action → editor on pharmacy-keyword activity (explicit check inside)
    """
    queryset = DrugDistribution.objects.prefetch_related(
        'items__stock__drug'
    ).select_related('beneficiary').all()
    serializer_class    = DrugDistributionSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = ['pharma', 'صيدل', 'drug', 'دواء', 'medicine']

    @action(detail=True, methods=['post'], url_path='validate')
    def validate_distribution(self, request, pk=None):
        """
        POST /api/distributions/{id}/validate/
        Validates the distribution and deducts stock atomically.
        Requires at least editor access.
        """
        # Extra explicit check for this sensitive action
        if not request.user.is_superuser:
            from .models import get_user_activities
            LEVEL_RANK = {'viewer': 1, 'editor': 2, 'manager': 3}
            accesses = get_user_activities(request.user)
            has_editor = any(
                any(kw.lower() in (a.activity.name or '').lower() for kw in self.activity_keywords)
                and LEVEL_RANK.get(a.access_level, 0) >= LEVEL_RANK['editor']
                for a in accesses
            )
            if not has_editor:
                return Response(
                    {'detail': 'You need editor access to validate distributions.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        distribution = self.get_object()
        try:
            distribution.validate()
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(DrugDistributionSerializer(distribution).data)


# ─────────────────────────────────────────────
# DASHBOARD
# Any authenticated user sees it.
# Numbers are scoped to what the user can access.
# ─────────────────────────────────────────────

class DashboardView(APIView):
    """
    GET /api/dashboard/
    Returns counts. Superuser sees everything.
    Regular users see counts scoped to their assigned activities.
    """
    permission_classes = [IsAuthenticatedAnyActivity]

    def get(self, request):
        user = request.user

        if user.is_superuser:
            # Admin sees global totals
            data = {
                'persons':        Person.objects.count(),
                'departments':    Department.objects.count(),
                'drugs':          Drug.objects.count(),
                'donations':      DrugDonation.objects.count(),
                'distributions':  DrugDistribution.objects.count(),
                'stock_entries':  DrugStock.objects.filter(quantity_available__gt=0).count(),
            }
        else:
            # Regular user sees only what their activities cover
            accesses = get_user_activities(user)
            data = {
                'your_activities': [
                    {
                        'activity':   a.activity.name,
                        'department': a.activity.department.name,
                        'level':      a.access_level,
                    }
                    for a in accesses
                ],
                'persons':       Person.objects.count(),
                'stock_entries': DrugStock.objects.filter(quantity_available__gt=0).count(),
            }

        return Response(data)


class ExpiringDrugsView(APIView):
    """
    GET /api/drugs/expiring/?days=30
    Requires viewer access on Pharmacy activity.
    """
    permission_classes = [IsAuthenticatedAnyActivity]

    def get(self, request):
        days   = int(request.query_params.get('days', 30))
        cutoff = timezone.now().date() + timedelta(days=days)
        stocks = DrugStock.objects.filter(
            expiration_date__lte=cutoff,
            quantity_available__gt=0,
        ).select_related('drug')
        return Response(DrugStockSerializer(stocks, many=True).data)

# ─────────────────────────────────────────────
# USER MANAGEMENT (superuser only)
# ─────────────────────────────────────────────

from .models import User, UserProfile, UserActivityAccess
from .serializers import UserCreateSerializer, UserUpdateSerializer

class UserViewSet(viewsets.ModelViewSet):
    """
    Full user CRUD for superuser.
    GET    /api/users/          → list all users with their accesses
    POST   /api/users/          → create user
    GET    /api/users/{id}/     → retrieve one user
    PUT    /api/users/{id}/     → update user (password optional)
    DELETE /api/users/{id}/     → delete user
    POST   /api/users/{id}/set_person/    → link user to a Person
    POST   /api/users/{id}/set_access/    → add/update an activity access
    DELETE /api/users/{id}/remove_access/ → remove an activity access
    """
    permission_classes = [IsAdminOnly]

    def get_queryset(self):
        return User.objects.exclude(is_superuser=True).prefetch_related(
            'activity_accesses__activity__department',
            'profile__person',
        ).order_by('username')

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action in ('update', 'partial_update'):
            return UserUpdateSerializer
        return UserSerializer

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return Response(UserSerializer(instance).data)

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        return Response(UserSerializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='set_person')
    def set_person(self, request, pk=None):
        """Link or unlink a Person to this user account."""
        user = self.get_object()
        person_id = request.data.get('person_id')  # pass null to unlink
        profile, _ = UserProfile.objects.get_or_create(user=user)
        if person_id:
            try:
                from .models import Person
                person = Person.objects.get(id=person_id)
                existing = UserProfile.objects.filter(person=person).exclude(user=user).first()
                if existing:
                    return Response(
                        {'detail': f'هذا الشخص مرتبط بالفعل بالمستخدم {existing.user.username}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                profile.person = person
            except Person.DoesNotExist:
                return Response({'detail': 'الشخص غير موجود'}, status=status.HTTP_404_NOT_FOUND)
        else:
            profile.person = None
        profile.save()
        return Response(UserSerializer(user).data)

    @action(detail=True, methods=['post'], url_path='set_access')
    def set_access(self, request, pk=None):
        """Add or update an activity access for this user."""
        user = self.get_object()
        activity_id  = request.data.get('activity_id')
        access_level = request.data.get('access_level', 'viewer')
        if not activity_id:
            return Response({'detail': 'activity_id مطلوب'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            activity = Activity.objects.get(id=activity_id)
        except Activity.DoesNotExist:
            return Response({'detail': 'النشاط غير موجود'}, status=status.HTTP_404_NOT_FOUND)
        access, created = UserActivityAccess.objects.update_or_create(
            user=user, activity=activity,
            defaults={'access_level': access_level}
        )
        return Response(UserSerializer(user).data)

    @action(detail=True, methods=['delete'], url_path='remove_access/(?P<access_id>[^/.]+)')
    def remove_access(self, request, pk=None, access_id=None):
        """Remove a specific activity access row."""
        user = self.get_object()
        try:
            access = UserActivityAccess.objects.get(id=access_id, user=user)
            access.delete()
        except UserActivityAccess.DoesNotExist:
            return Response({'detail': 'الصلاحية غير موجودة'}, status=status.HTTP_404_NOT_FOUND)
        return Response(UserSerializer(user).data)
    

# ─────────────────────────────────────────────
# BLOOD DONATION ACTIVITY ("قطرة حياة")
# ─────────────────────────────────────────────

BLOOD_KEYWORDS = ['قطرة حياة', 'قطرة', 'دم', 'blood']


class DonorViewSet(viewsets.ModelViewSet):
    queryset            = Donor.objects.select_related('person').all()
    serializer_class    = DonorSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = BLOOD_KEYWORDS


class PatientViewSet(viewsets.ModelViewSet):
    queryset            = Patient.objects.select_related('person').all()
    serializer_class    = PatientSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = BLOOD_KEYWORDS


class DonationHistoryViewSet(viewsets.ModelViewSet):
    queryset = DonationHistory.objects.select_related('donor__person', 'patient__person').all()
    serializer_class    = DonationHistorySerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = BLOOD_KEYWORDS


class DonateBloodView(APIView):
    """
    POST /api/donate/<id_patient>/<id_donor>/
    Requires editor access on the blood-donation activity.
    """
    permission_classes = [HasActivityAccess]
    activity_keywords  = BLOOD_KEYWORDS

    def post(self, request, id_patient, id_donor):
        donor   = get_object_or_404(Donor.objects.select_related('person'), id=id_donor)
        patient = get_object_or_404(Patient.objects.select_related('person'), id=id_patient)

        # Your model already had can_donate()/is_approved — wiring them in here,
        # since the old donate_blood() never actually called them.
        if not donor.is_approved:
            return Response({'error': 'هذا المتبرع غير معتمد بعد.'}, status=status.HTTP_400_BAD_REQUEST)
        if not donor.can_donate:
            return Response(
                {'error': 'يجب الانتظار 90 يوماً على الأقل منذ آخر تبرع لهذا المتبرع.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            donation = DonationHistory.objects.create(donor=donor, patient=patient)
            donor.date_last_donation = date.today()
            donor.save()

        return Response(
            {'message': 'تم تسجيل التبرع بنجاح.', 'donation': DonationHistorySerializer(donation).data},
            status=status.HTTP_201_CREATED,
        )


class CompatibleDonorsView(APIView):
    """GET /api/patients/with-compatible-donors/"""
    permission_classes = [HasActivityAccess]
    activity_keywords  = BLOOD_KEYWORDS

    def get(self, request):
        three_months_ago = date.today() - timedelta(days=90)
        result = []
        for patient in Patient.objects.filter(is_active=True).select_related('person'):
            donors = Donor.objects.filter(
                blood_type=patient.blood_type, is_approved=True
            ).filter(
                Q(date_last_donation__isnull=True) | Q(date_last_donation__lte=three_months_ago)
            ).select_related('person')
            result.append({
                'patient': PatientSerializer(patient).data,
                'donors':  DonorSerializer(donors, many=True).data,
            })
        return Response(result)


class BloodDonationDashboardView(APIView):
    """
    GET /api/dashboard/stats/
    so I wrote it fresh — adjust the fields to whatever your dashboard actually needs.
    """
    permission_classes = [HasActivityAccess]
    activity_keywords  = BLOOD_KEYWORDS

    def get(self, request):
        three_months_ago = date.today() - timedelta(days=90)
        eligible_now = Donor.objects.filter(is_approved=True).filter(
            Q(date_last_donation__isnull=True) | Q(date_last_donation__lte=three_months_ago)
        ).count()
        return Response({
            'donors_total':           Donor.objects.count(),
            'donors_approved':        Donor.objects.filter(is_approved=True).count(),
            'donors_eligible_now':    eligible_now,
            'patients_total':         Patient.objects.count(),
            'patients_active':        Patient.objects.filter(is_active=True).count(),
            'donations_total':        DonationHistory.objects.count(),
            'donations_last_30_days': DonationHistory.objects.filter(
                donation_date__gte=date.today() - timedelta(days=30)
            ).count(),
        })


try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    ARABIC_SUPPORT = True
except ImportError:
    ARABIC_SUPPORT = False
    print("Warning: arabic-reshaper or python-bidi not installed.")


def reshape_arabic(text):
    """تحويل النص العربي للعرض الصحيح في PDF"""
    if not text:
        return ""
    if ARABIC_SUPPORT:
        try:
            reshaped_text = arabic_reshaper.reshape(str(text))
            bidi_text = get_display(reshaped_text)
            return bidi_text
        except:
            return str(text)
    else:
        return str(text)

class CertificateView(APIView):
    """
    GET /api/certificate/<patient_id>/<donor_id>/
    Requires: pip install reportlab arabic-reshaper python-bidi
    Requires a logo file at <BASE_DIR>/static/images/logo.jpg (or .png).
    """
    permission_classes = [HasActivityAccess]
    activity_keywords  = BLOOD_KEYWORDS

    def get(self, request, patient_id, donor_id):
        """توليد شهادة تبرع بالدم بتصميم احترافي أبيض وأسود"""
        patient = get_object_or_404(Patient, id=patient_id)
        donor = get_object_or_404(Donor, id=donor_id)

        response = HttpResponse(content_type='application/pdf')
        filename = f'certificate_{patient_id}_{donor_id}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4

        arabic_font = 'Helvetica'
        try:
            possible_font_paths = [
                os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf'),
                os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NotoSansArabic-Regular.ttf'),
                os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Arial.ttf'),
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                'C:\\Windows\\Fonts\\arial.ttf',
            ]
            for font_path in possible_font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
        except Exception as e:
            print(f"Font error: {e}")

        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.jpg')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.jpg')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png')

        if os.path.exists(logo_path):
            try:
                logo = ImageReader(logo_path)
                logo_width = 3 * cm
                logo_height = 3 * cm
                p.drawImage(logo, width / 2 - logo_width / 2, height - 5.25 * cm,
                            width=logo_width, height=logo_height, mask='auto')
            except Exception as e:
                print(f"Logo error: {e}")
        else:
            print(f"Logo not found at: {logo_path}")

        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(3)
        p.rect(1.5 * cm, 1.5 * cm, width - 3 * cm, height - 3 * cm, stroke=1, fill=0)
        p.setLineWidth(1)
        p.rect(2 * cm, 2 * cm, width - 4 * cm, height - 4 * cm, stroke=1, fill=0)

        corner_size = 1 * cm
        p.line(width - 2 * cm, height - 2 * cm, width - 2 * cm - corner_size, height - 2 * cm)
        p.line(width - 2 * cm, height - 2 * cm, width - 2 * cm, height - 2 * cm - corner_size)
        p.line(2 * cm, height - 2 * cm, 2 * cm + corner_size, height - 2 * cm)
        p.line(2 * cm, height - 2 * cm, 2 * cm, height - 2 * cm - corner_size)
        p.line(width - 2 * cm, 2 * cm, width - 2 * cm - corner_size, 2 * cm)
        p.line(width - 2 * cm, 2 * cm, width - 2 * cm, 2 * cm + corner_size)
        p.line(2 * cm, 2 * cm, 2 * cm + corner_size, 2 * cm)
        p.line(2 * cm, 2 * cm, 2 * cm, 2 * cm + corner_size)

        p.setFont(arabic_font, 26)
        p.setFillColorRGB(0, 0, 0)
        title = reshape_arabic("جمعية الغد الأفضل")
        p.drawCentredString(width / 2, height - 6 * cm, title)

        p.setLineWidth(2)
        p.line(width / 2 - 4 * cm, height - 6.5 * cm, width / 2 + 4 * cm, height - 6.5 * cm)

        p.setFont(arabic_font, 18)
        p.setFillColorRGB(0.2, 0.2, 0.2)
        subtitle = reshape_arabic("استمارة التبرع بالدم")
        p.drawCentredString(width / 2, height - 7.5 * cm, subtitle)

        y_position = height - 9.5 * cm
        box_height = 7 * cm
        p.setFillColorRGB(0.95, 0.95, 0.95)
        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(1)
        p.rect(3 * cm, y_position - box_height, width - 6 * cm, box_height, stroke=1, fill=1)

        header_height = 0.8 * cm
        p.setFillColorRGB(0, 0, 0)
        p.rect(3 * cm, y_position - header_height, width - 6 * cm, header_height, stroke=0, fill=1)
        p.setFont(arabic_font, 14)
        p.setFillColorRGB(1, 1, 1)
        p.drawCentredString(width / 2, y_position - 0.55 * cm, reshape_arabic("معلومات المتبرع"))

        y_position -= 1.5 * cm
        p.setFont(arabic_font, 11)
        p.setFillColorRGB(0, 0, 0)
        right_margin = width - 5 * cm
        label_x = right_margin
        value_x = right_margin - 3 * cm

        def field(label, value, size=11):
            nonlocal y_position
            p.setFont(arabic_font, 10)
            p.setFillColorRGB(0.3, 0.3, 0.3)
            p.drawRightString(label_x, y_position, reshape_arabic(label))
            p.setFont(arabic_font, size)
            p.setFillColorRGB(0, 0, 0)
            p.drawRightString(value_x, y_position, value if value and label == "الزمرة الدموية:" else reshape_arabic(value or ''))
            y_position -= 0.7 * cm

        field("الاسم:", donor.person.first_name)
        field("اللقب:", donor.person.last_name)
        field("تاريخ الميلاد:", donor.person.date_of_birth.strftime('%d-%m-%Y') if donor.person.date_of_birth else 'غير محدد')
        field("الزمرة الدموية:", donor.blood_type or '', size=13)
        field("تاريخ آخر تبرع:", donor.date_last_donation.strftime('%d-%m-%Y') if donor.date_last_donation else 'أول تبرع')

        p.setFont(arabic_font, 10)
        p.setFillColorRGB(0.3, 0.3, 0.3)
        p.drawRightString(label_x, y_position, reshape_arabic("معلومات إضافية:"))
        p.setFont(arabic_font, 9)
        p.setFillColorRGB(0, 0, 0)
        description = donor.description or 'لا توجد معلومات إضافية'
        if len(description) > 30:
            words = description.split()
            lines, current_line = [], ""
            for word in words:
                if len(current_line + " " + word) <= 30:
                    current_line += " " + word if current_line else word
                else:
                    lines.append(current_line)
                    current_line = word
            if current_line:
                lines.append(current_line)
            for i, line in enumerate(lines[:2]):
                p.drawRightString(value_x, y_position - (i * 0.5 * cm), reshape_arabic(line))
        else:
            p.drawRightString(value_x, y_position, reshape_arabic(description))

        y_position -= 2.5 * cm
        box_height = 5.8 * cm
        p.setFillColorRGB(0.95, 0.95, 0.95)
        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(1)
        p.rect(3 * cm, y_position - box_height, width - 6 * cm, box_height, stroke=1, fill=1)
        p.setFillColorRGB(0, 0, 0)
        p.rect(3 * cm, y_position - header_height, width - 6 * cm, header_height, stroke=0, fill=1)
        p.setFont(arabic_font, 14)
        p.setFillColorRGB(1, 1, 1)
        p.drawCentredString(width / 2, y_position - 0.55 * cm, reshape_arabic("معلومات المريض"))

        y_position -= 1.5 * cm
        field("الاسم:", patient.person.first_name)
        field("اللقب:", patient.person.last_name)
        field("تاريخ الميلاد:", patient.person.date_of_birth.strftime('%d-%m-%Y') if patient.person.date_of_birth else 'غير محدد')
        field("الزمرة الدموية:", patient.blood_type or '', size=13)
        field("المستشفى:", patient.hospital_name or 'غير محدد')

        y_position -= 1.5 * cm
        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(1.5)
        p.line(4 * cm, y_position, width - 4 * cm, y_position)

        y_position = 5.5 * cm
        p.setFont(arabic_font, 12)
        p.setFillColorRGB(0, 0, 0)
        p.drawString(4.5 * cm, y_position, reshape_arabic("رئيس الجمعية"))
        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(0.8)
        p.line(4 * cm, y_position - 0.3 * cm, 7 * cm, y_position - 0.3 * cm)

        p.setFont(arabic_font, 8)
        p.setFillColorRGB(0.4, 0.4, 0.4)
        footer = reshape_arabic("جمعية الغد الأفضل - نساهم في إنقاذ الأرواح")
        p.drawCentredString(width / 2, 2.5 * cm, footer)
        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(0.5)
        p.line(3 * cm, 3 * cm, width - 3 * cm, 3 * cm)

        p.showPage()
        p.save()
        return response
    

# machine


# Keywords that match your machines activity in the admin panel
MACHINE_KEYWORDS = ['machine', 'جهاز', 'equipment', 'معدات']


def normalize_machine_barcode(raw_code):
    """
    Normalize barcodes generated as M-{PREFIX}{id}, e.g. 'm-ct05' and
    'M-CT5' should both resolve to 'M-CT5' — strips leading zeros from the
    trailing numeric id, uppercases the rest, and preserves the '-' separator.
    """
    code = (raw_code or '').strip().upper()
    if code.startswith('M-'):
        body = code[2:]
        # split into the alpha prefix and the trailing digit run
        i = len(body)
        while i > 0 and body[i - 1].isdigit():
            i -= 1
        prefix, digits = body[:i], body[i:]
        if digits.isdigit():
            return f"M-{prefix}{int(digits)}"
    return code

class MachineViewSet(viewsets.ModelViewSet):
    """
    GET            → viewer
    POST/PUT/PATCH → editor
    DELETE         → manager
    Requires an activity whose name contains one of MACHINE_KEYWORDS.
    """
    queryset           = Machine.objects.all() 
    serializer_class   = MachineSerializer
    permission_classes = [HasActivityAccess]
    activity_keywords  = MACHINE_KEYWORDS

    def get_serializer_context(self):
        # Pass request so get_photo_url can build absolute URLs
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    @action(detail=False, methods=['get'], url_path='by-barcode/(?P<bar_code>[^/.]+)')
    def by_barcode(self, request, bar_code=None):
        """GET /api/machines/by-barcode/<bar_code>/"""
        machine = get_object_or_404(Machine, bar_code=normalize_machine_barcode(bar_code))
        return Response(MachineSerializer(machine, context={'request': request}).data)

    @action(detail=True, methods=['patch'], url_path='set-status')
    def set_status(self, request, pk=None):
        """
        PATCH /api/machines/{id}/set-status/
        Body: { "status": "maintenance" }
        Editor level required. 'assigned' is blocked — use assignments instead.
        """
        machine    = self.get_object()
        new_status = request.data.get('status')
        allowed    = ['available', 'destroyed', 'maintenance']
        if new_status not in allowed:
            return Response(
                {'detail': f'القيم المسموح بها: {allowed}'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        machine.status = new_status
        machine.save()
        return Response(MachineSerializer(machine, context={'request': request}).data)
    @action(detail=False, methods=["post"], url_path="print-barcodes")
    def print_barcodes(self, request):
        """
        POST /api/machines/print-barcodes/
        Body: { "ids": [1, 2, 3, ...] }
        Returns a multi-page PDF with barcode labels for each valid machine.
        Missing/invalid IDs are reported via the X-Missing-Ids response header.
        """
        ids = request.data.get("ids", [])
        if not ids:
            return Response(
                {"detail": "No machines selected."},
                status=status.HTTP_400_BAD_REQUEST
            )
    
        machines = Machine.objects.filter(id__in=ids)
        if not machines.exists():
            return Response(
                {"detail": "No valid machines found for the provided IDs."},
                status=status.HTTP_404_NOT_FOUND
            )
    
        # Track which requested ids weren't found
        found_ids = set(machines.values_list("id", flat=True))
        missing_ids = [str(i) for i in ids if int(i) not in found_ids]
    
        # PDF generation
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
    
        # Layout settings — 3 labels per row, bordered cards
        labels_per_row = 3
        margin_x = 30
        margin_y = 30
        x_gap = 15
        y_gap = 15
    
        label_width = (width - 2 * margin_x - (labels_per_row - 1) * x_gap) / labels_per_row
        label_height = 105
        barcode_height = 45
        title_font_size = 12
        code_font_size = 10
    
        # Optional page title
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(width / 2, height - margin_y + 5, "Machine Barcodes")
    
        x = margin_x
        y = height - margin_y - 25 - label_height  # leave room for page title on first page
    
        def draw_label(c, x, y, machine):
            """Draws one bordered label card at top-left corner (x, y)."""
            local_title_font_size = title_font_size  # local copy, doesn't touch outer scope
        
            # Border box
            c.setStrokeColorRGB(0, 0, 0)
            c.setLineWidth(1)
            c.rect(x, y, label_width, label_height, stroke=1, fill=0)
        
            # Title (machine name), centered near the top
            title = machine.name
            while c.stringWidth(title, "Helvetica-Bold", local_title_font_size) > label_width - 10 and local_title_font_size > 7:
                local_title_font_size -= 1
            c.setFont("Helvetica-Bold", local_title_font_size)
            c.drawCentredString(x + label_width / 2, y + label_height - 20, title)
        
            # Barcode image, centered horizontally
            CODE128 = barcode.get_barcode_class('code128')
            code = CODE128(machine.bar_code or 'N/A', writer=ImageWriter())
            barcode_buffer = BytesIO()
            code.write(barcode_buffer, options={"write_text": False})
            barcode_buffer.seek(0)
            barcode_image = ImageReader(barcode_buffer)
        
            bc_width = label_width - 20
            bc_x = x + (label_width - bc_width) / 2
            bc_y = y + (label_height - barcode_height) / 2 - 5
            c.drawImage(
                barcode_image, bc_x, bc_y,
                width=bc_width, height=barcode_height,
                preserveAspectRatio=False, mask='auto'
            )

            # Bar code text, centered below the barcode
            c.setFont("Helvetica", code_font_size)
            c.drawCentredString(
                x + label_width / 2,
                bc_y - 12,
                machine.bar_code or 'N/A'
            )
    
        first_page = True
        for machine in machines:
            draw_label(c, x, y, machine)
    
            # Move to next horizontal slot
            x += label_width + x_gap
            if x + label_width > width - margin_x + 1:
                x = margin_x
                y -= label_height + y_gap
    
                # New page if we've run out of vertical space
                if y < margin_y:
                    c.showPage()
                    first_page = False
                    x = margin_x
                    y = height - margin_y - label_height
    
        c.save()
        buffer.seek(0)
        pdf = buffer.getvalue()
    
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="machine_barcodes.pdf"'
        if missing_ids:
            response["X-Missing-Ids"] = ",".join(missing_ids)
        return response

    @action(detail=True, methods=['get'], url_path='history')
    def history(self, request, pk=None):
        """GET /api/machines/{id}/history/  → all assignments for this machine"""
        machine     = self.get_object()
        assignments = machine.assignments.select_related('assigned_to').all()
        return Response(MachineAssignmentSerializer(assignments, many=True).data)


class MachineAssignmentViewSet(viewsets.ModelViewSet):
    """
    GET    → viewer
    POST   → editor  (assign machine)
    DELETE → manager (delete record entirely — prefer return_machine instead)

    POST /api/machine-assignments/{id}/return/  → mark as returned
    """
    queryset           = MachineAssignment.objects.select_related('machine', 'assigned_to').all()
    serializer_class   = MachineAssignmentSerializer
    permission_classes = [HasActivityAccess]
    activity_keywords  = MACHINE_KEYWORDS

    @staticmethod
    def _normalize_bar_code(raw_code):
        """Accept GA1 and GA01 as the same machine barcode."""
        return normalize_machine_barcode(raw_code)
    def destroy(self, request, *args, **kwargs):
        return Response(
            {'detail': 'لا يمكن حذف سجل الإسناد نهائياً.'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if timezone.now() - instance.assigned_at > timedelta(days=1):
            return Response(
                {'detail': 'لا يمكن تعديل هذا الإسناد بعد مرور أكثر من يوم واحد على تاريخ الإسناد.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        if timezone.now() - instance.assigned_at > timedelta(days=1):
            return Response(
                {'detail': 'لا يمكن تعديل هذا الإسناد بعد مرور أكثر من يوم واحد على تاريخ الإسناد.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().partial_update(request, *args, **kwargs)
    def get_queryset(self):
        qs        = super().get_queryset()
        active    = self.request.query_params.get('active')     # ?active=true
        machine   = self.request.query_params.get('machine')    # ?machine=5
        person    = self.request.query_params.get('person')     # ?person=3
        if active == 'true':
            qs = qs.filter(returned_at__isnull=True)
        if machine:
            qs = qs.filter(machine_id=machine)
        if person:
            qs = qs.filter(assigned_to_id=person)
        return qs

    @action(detail=False, methods=['post'], url_path='assign-by-barcode')
    def assign_by_barcode(self, request):
        """
        POST /api/machine-assignments/assign-by-barcode/
        Body: { "bar_code": "GA01", "assigned_to": 3, "description": "..." }
        """
        bar_code = self._normalize_bar_code(request.data.get('bar_code'))
        assigned_to = request.data.get('assigned_to')
        description = request.data.get('description', '')
        assigned_at_raw = request.data.get('assigned_at')
        if not bar_code or not assigned_to:
            return Response(
                {'detail': 'bar_code و assigned_to مطلوبان.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        machine = Machine.objects.filter(bar_code=bar_code).first()
        if not machine:
            return Response(
                {'detail': f'لم يتم العثور على جهاز بالباركود {bar_code}.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if machine.status != 'available':
            return Response(
                {'detail': 'هذا الجهاز غير متاح حالياً للإسناد.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        data={
            'machine': machine.id,
            'assigned_to': assigned_to,
            'description': description,
        }
        if assigned_at_raw:  # NEW — only include if user supplied one
          data['assigned_at'] = assigned_at_raw

        serializer = self.get_serializer(data)
        serializer.is_valid(raise_exception=True)
        assignment = serializer.save()
        return Response(
            self.get_serializer(assignment).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'], url_path='return-by-barcode')
    def return_by_barcode(self, request):
        """
        POST /api/machine-assignments/return-by-barcode/
        Body: { "bar_code": "GA01", "returned_at": "2026-06-29T10:00:00Z", "description": "..." }
        """
        bar_code = self._normalize_bar_code(request.data.get('bar_code'))
        returned_at_raw = request.data.get('returned_at')
        return_description = request.data.get('description', '')

        if not bar_code:
            return Response({'detail': 'bar_code مطلوب.'}, status=status.HTTP_400_BAD_REQUEST)

        machine = Machine.objects.filter(bar_code=bar_code).first()
        if not machine:
            return Response(
                {'detail': f'لم يتم العثور على جهاز بالباركود {bar_code}.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        assignment = (
            MachineAssignment.objects
            .filter(machine=machine, returned_at__isnull=True)
            .order_by('-assigned_at')
            .first()
        )
        if not assignment:
            return Response(
                {'detail': 'لا يوجد إسناد نشط لهذا الجهاز.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        returned_at = None
        if returned_at_raw:
            returned_at = parse_datetime(returned_at_raw)
            if returned_at is None:
                return Response(
                    {'detail': 'صيغة returned_at غير صحيحة. استخدم ISO datetime.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if timezone.is_naive(returned_at):
                returned_at = timezone.make_aware(returned_at, timezone.get_current_timezone())

        try:
            assignment.return_machine(returned_at=returned_at, return_description=return_description)
        except ValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(self.get_serializer(assignment).data)

    @action(detail=True, methods=['post'], url_path='return')
    def return_machine(self, request, pk=None):
        """
        POST /api/machine-assignments/{id}/return/
        Optional body: { "returned_at": "ISO datetime", "description": "..." }
        """
        assignment = self.get_object()
        returned_at_raw = request.data.get('returned_at')
        return_description = request.data.get('description', '')

        returned_at = None
        if returned_at_raw:
            returned_at = parse_datetime(returned_at_raw)
            if returned_at is None:
                return Response(
                    {'detail': 'صيغة returned_at غير صحيحة. استخدم ISO datetime.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if timezone.is_naive(returned_at):
                returned_at = timezone.make_aware(returned_at, timezone.get_current_timezone())

        try:
            assignment.return_machine(returned_at=returned_at, return_description=return_description)
        except ValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(MachineAssignmentSerializer(assignment).data)
    


FINANCE_KEYWORDS = ['مالية', 'finance', 'مال', 'ميزانية']


class FinancialCategoryViewSet(viewsets.ModelViewSet):
    """
    GET /api/finance/categories/             → all categories (management page)
    GET /api/finance/categories/?active=true → active categories only (dashboard)
    """
    queryset           = FinancialCategory.objects.all()
    serializer_class   = FinancialCategorySerializer
    permission_classes = [HasActivityAccess]
    activity_keywords  = FINANCE_KEYWORDS

    def get_queryset(self):
        qs = super().get_queryset()
        active = self.request.query_params.get('active')
        if active is not None:
            active_bool = str(active).lower() in ('1', 'true', 'yes')
            qs = qs.filter(is_active=active_bool)
        return qs.order_by('-is_active', 'name')


class DonationViewSet(viewsets.ModelViewSet):
    queryset            = Donation.objects.select_related('category', 'created_by').all()
    serializer_class    = DonationSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = FINANCE_KEYWORDS

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    def get_queryset(self):
        qs = super().get_queryset()
        donor = self.request.query_params.get('donor')
        category = self.request.query_params.get('category')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if donor:
            qs = qs.filter(donor_name__icontains=donor)
        if category:
            qs = qs.filter(category_id=category)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        return qs

    def perform_destroy(self, instance):
        from .models import log_finance_action
        log_finance_action(self.request.user, 'delete', 'Donation', instance.id,
                            f"{instance.donor_name} - {instance.amount} DZD")
        instance.delete()

    @action(detail=False, methods=['post'], url_path='print-receipts')
    def print_receipts(self, request):
        """
        POST /api/finance/donations/print-receipts/
        Body: { "ids": [1, 2, 3, ...] }
        Returns a multi-page PDF with 4 donation receipts per A4 page (2x2 grid),
        same pattern as MachineViewSet.print_barcodes.
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'No donations selected.'}, status=status.HTTP_400_BAD_REQUEST)

        donations = Donation.objects.filter(id__in=ids).select_related('category')
        if not donations.exists():
            return Response({'detail': 'No valid donations found for the provided IDs.'},
                             status=status.HTTP_404_NOT_FOUND)

        found_ids = set(donations.values_list('id', flat=True))
        missing_ids = [str(i) for i in ids if int(i) not in found_ids]

        items = [
            _donation_receipt_item(d) for d in donations
        ]
        response = generate_receipts_pdf(items, filename='donation_receipts.pdf')
        if missing_ids:
            response['X-Missing-Ids'] = ','.join(missing_ids)
        return response


class ExpenseTransactionViewSet(viewsets.ModelViewSet):
    queryset            = ExpenseTransaction.objects.select_related('category', 'created_by', 'related_donation').all()
    serializer_class    = ExpenseTransactionSerializer
    permission_classes  = [HasActivityAccess]
    activity_keywords   = FINANCE_KEYWORDS

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    def get_queryset(self):
        qs = super().get_queryset()
        category = self.request.query_params.get('category')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if category:
            qs = qs.filter(category_id=category)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        return qs

    def perform_destroy(self, instance):
        from .models import log_finance_action
        log_finance_action(self.request.user, 'delete', 'ExpenseTransaction', instance.id,
                            f"{instance.amount} DZD - {instance.description}")
        instance.delete()

    @action(detail=False, methods=['post'], url_path='print-receipts')
    def print_receipts(self, request):
        """
        POST /api/finance/expenses/print-receipts/
        Body: { "ids": [1, 2, 3, ...] }
        Returns a multi-page PDF with 4 expense receipts per A4 page (2x2 grid).
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'No expenses selected.'}, status=status.HTTP_400_BAD_REQUEST)

        expenses = ExpenseTransaction.objects.filter(id__in=ids).select_related('category')
        if not expenses.exists():
            return Response({'detail': 'No valid expenses found for the provided IDs.'},
                             status=status.HTTP_404_NOT_FOUND)

        found_ids = set(expenses.values_list('id', flat=True))
        missing_ids = [str(i) for i in ids if int(i) not in found_ids]

        items = [_expense_receipt_item(e) for e in expenses]
        response = generate_receipts_pdf(items, filename='expense_receipts.pdf')
        if missing_ids:
            response['X-Missing-Ids'] = ','.join(missing_ids)
        return response


class FinancialDashboardView(APIView):
    """GET /api/finance/dashboard/"""
    permission_classes = [HasActivityAccess]
    activity_keywords  = FINANCE_KEYWORDS

    def get(self, request):
        total_donations = Donation.objects.aggregate(total=Sum('amount'))['total'] or 0
        total_expenses  = ExpenseTransaction.objects.aggregate(total=Sum('amount'))['total'] or 0
        balance = total_donations - total_expenses

        today = date.today()
        month_donations = Donation.objects.filter(
            date__year=today.year, date__month=today.month
        ).aggregate(total=Sum('amount'))['total'] or 0
        month_expenses = ExpenseTransaction.objects.filter(
            date__year=today.year, date__month=today.month
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Only active categories are surfaced on the dashboard. Inactive
        # categories stay visible/manageable on the Categories page but are
        # excluded here per the "Active Categories on Dashboard" requirement.
        by_category = list(
            ExpenseTransaction.objects.filter(category__is_active=True)
            .values('category__name')
            .annotate(total=Sum('amount'))
            .order_by('-total')
        )

        return Response({
            'balance': balance,
            'total_donations': total_donations,
            'total_expenses': total_expenses,
            'month_donations': month_donations,
            'month_expenses': month_expenses,
            'spending_by_category': by_category,
        })
    # PATCH method removed entirely — no more threshold setting


def _donation_receipt_item(donation):
    return dict(
        title='إيصال تبرع',
        name_label='المتبرع', name_value=donation.donor_name,
        amount=donation.amount,
        date_value=donation.date,
        extra_label='الفئة', extra_value=donation.category.name,
        method_label='طريقة الدفع',
        method_value='نقداً' if donation.payment_method == 'cash' else 'حساب بنكي',
        ref=f"D-{donation.id:06d}",
    )


def _expense_receipt_item(expense):
    return dict(
        title='إيصال الدفع',
        name_label='الوصف', name_value=expense.description,
        amount=expense.amount,
        date_value=expense.date,
        extra_label='الفئة', extra_value=expense.category.name,
        method_label=None, method_value=None,
        ref=f"E-{expense.id:06d}",
    )


class DonationReceiptView(APIView):
    """
    GET /api/finance/donations/{id}/receipt/
    Generates an A4 PDF laid out with the 4-up receipt grid (see
    generate_receipts_pdf), with this single receipt in the first cell.
    """
    permission_classes = [HasActivityAccess]
    activity_keywords  = FINANCE_KEYWORDS

    def get(self, request, pk):
        donation = get_object_or_404(Donation, pk=pk)
        return generate_receipts_pdf([_donation_receipt_item(donation)], filename='donation_receipt.pdf')


class ExpenseReceiptView(APIView):
    """GET /api/finance/expenses/{id}/receipt/"""
    permission_classes = [HasActivityAccess]
    activity_keywords  = FINANCE_KEYWORDS

    def get(self, request, pk):
        expense = get_object_or_404(ExpenseTransaction, pk=pk)
        return generate_receipts_pdf([_expense_receipt_item(expense)], filename='expense_receipt.pdf')
class FinancialReportView(APIView):
    """
    GET /api/finance/reports/?year=2026&month=7&category=3&format=json|excel|pdf
    """
    permission_classes = [HasActivityAccess]
    activity_keywords  = FINANCE_KEYWORDS

    def get(self, request):
        donations = Donation.objects.select_related('category').all()
        expenses  = ExpenseTransaction.objects.select_related('category').all()

        year = request.query_params.get('year')
        month = request.query_params.get('month')
        category = request.query_params.get('category')  # comma-separated ids, e.g. "1,3,5"

        if year:
            donations = donations.filter(date__year=year)
            expenses  = expenses.filter(date__year=year)
        if month:
            donations = donations.filter(date__month=month)
            expenses  = expenses.filter(date__month=month)

        category_ids = [c for c in (category.split(',') if category else []) if c]
        if category_ids:
            donations = donations.filter(category_id__in=category_ids)
            expenses  = expenses.filter(category_id__in=category_ids)

        fmt = request.query_params.get('export_format', 'json')

        categories_label = None
        if category_ids:
            cat_names = list(
                FinancialCategory.objects.filter(id__in=category_ids).values_list('name', flat=True)
            )
            categories_label = '، '.join(sorted(cat_names)) if cat_names else None

        if fmt == 'excel':
            return self._export_excel(donations, expenses, year, month, categories_label)
        if fmt == 'pdf':
            return self._export_pdf(donations, expenses, year, month, categories_label)

        total_donations = donations.aggregate(total=Sum('amount'))['total'] or 0
        total_expenses  = expenses.aggregate(total=Sum('amount'))['total'] or 0
        return Response({
            'total_donations': total_donations,
            'total_expenses': total_expenses,
            'net': total_donations - total_expenses,
            'donations': DonationSerializer(donations, many=True, context={'request': request}).data,
            'expenses': ExpenseTransactionSerializer(expenses, many=True, context={'request': request}).data,
        })

    def _report_meta(self, donations, expenses, year, month, categories_label=None):
        if categories_label:
            categories_label_final = categories_label
        else:
            names = sorted(set(
                list(donations.values_list('category__name', flat=True)) +
                list(expenses.values_list('category__name', flat=True))
            ))
            categories_label_final = '، '.join(names) if names else 'كل الفئات'

        MONTH_NAMES_AR = ['', 'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                          'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
        period_bits = []
        if year:
            period_bits.append(str(year))
        if month:
            try:
                period_bits.append(MONTH_NAMES_AR[int(month)])
            except (ValueError, IndexError):
                period_bits.append(str(month))
        period_label = ' - '.join(period_bits) if period_bits else 'كل الفترات'

        return {
            'title': 'التقرير المالي',
            'description': 'تقرير مالي شامل يوضح التبرعات والمصروفات المسجلة لدى الجمعية خلال الفترة المحددة.',
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
            'year': str(year) if year else 'كل السنوات',
            'period': period_label,
            'categories': categories_label_final,
        }

    def _export_excel(self, donations, expenses, year=None, month=None, categories_label=None):
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        meta = self._report_meta(donations, expenses, year, month, categories_label)
        header_fill = PatternFill(start_color='14B8A6', end_color='14B8A6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        title_font = Font(bold=True, size=14)

        wb = Workbook()

        # ── Summary sheet ────────────────────────────────────────────────
        total_donations = sum(float(d.amount) for d in donations)
        total_expenses  = sum(float(e.amount) for e in expenses)
        ws0 = wb.active
        ws0.title = 'ملخص'
        ws0['A1'] = ORG_NAME
        ws0['A1'].font = title_font
        ws0['A2'] = meta['title']
        ws0['A2'].font = Font(bold=True, size=12)
        ws0['A3'] = f"تاريخ الإصدار: {meta['generated_at']}"
        ws0['A4'] = f"الفترة: {meta['period']}"
        ws0['A5'] = f"الفئات: {meta['categories']}"
        ws0['A7'] = 'إجمالي التبرعات'
        ws0['B7'] = total_donations
        ws0['A8'] = 'إجمالي المصروفات'
        ws0['B8'] = total_expenses
        ws0['A9'] = 'الصافي'
        ws0['B9'] = total_donations - total_expenses
        for r in (7, 8, 9):
            ws0[f'A{r}'].font = Font(bold=True)
        ws0.column_dimensions['A'].width = 28
        ws0.column_dimensions['B'].width = 18

        # ── Donations sheet ──────────────────────────────────────────────
        ws1 = wb.create_sheet('التبرعات')
        headers1 = ['المتبرع', 'المبلغ', 'طريقة الدفع', 'الفئة', 'التاريخ', 'ملاحظات']
        ws1.append(headers1)
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for d in donations:
            ws1.append([d.donor_name, float(d.amount), d.get_payment_method_display(),
                        d.category.name, str(d.date), d.notes or ''])
        for i, w in enumerate([22, 14, 16, 20, 14, 30], start=1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.freeze_panes = 'A2'

        # ── Expenses sheet ───────────────────────────────────────────────
        ws2 = wb.create_sheet('المصروفات')
        headers2 = ['المبلغ', 'الفئة', 'الوصف', 'التاريخ']
        ws2.append(headers2)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for e in expenses:
            ws2.append([float(e.amount), e.category.name, e.description, str(e.date)])
        for i, w in enumerate([14, 20, 40, 14], start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="financial_report.xlsx"'
        wb.save(response)
        return response

    def _export_pdf(self, donations, expenses, year=None, month=None, categories_label=None):
        width, height = A4
        margin = 2 * cm
        arabic_font = _get_arabic_font()
        rs = _get_reshaper()
        logo_path = _find_logo_path()
        meta = self._report_meta(donations, expenses, year, month, categories_label)

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        page_num = [1]

        def draw_footer():
            c.setFont(arabic_font, 8)
            c.setFillColorRGB(0.45, 0.45, 0.45)
            c.drawCentredString(width / 2, 1 * cm, rs(ORG_NAME))
            c.drawRightString(width - margin, 1 * cm, rs(f"صفحة {page_num[0]}"))
            c.setFillColorRGB(0, 0, 0)

        def draw_header(is_first_page):
            y = height - margin
            if is_first_page and logo_path:
                try:
                    logo_w = logo_h = 2 * cm
                    logo = ImageReader(logo_path)
                    c.drawImage(logo, width / 2 - logo_w / 2, y - logo_h,
                                width=logo_w, height=logo_h, mask='auto', preserveAspectRatio=True)
                except Exception:
                    pass
                y -= logo_h + 0.3 * cm

            c.setFont(arabic_font, 15)
            c.drawCentredString(width / 2, y, rs(ORG_NAME)); y -= 0.65 * cm
            c.setFont(arabic_font, 13)
            c.drawCentredString(width / 2, y, rs(meta['title'])); y -= 0.55 * cm

            if is_first_page:
                c.setFont(arabic_font, 8.5)
                c.drawCentredString(width / 2, y, rs(meta['description'])); y -= 0.55 * cm

                c.setLineWidth(0.8)
                c.setStrokeColorRGB(0.75, 0.75, 0.75)
                c.line(margin, y, width - margin, y); y -= 0.5 * cm

                c.setFont(arabic_font, 9)
                for line in (
                    f"تاريخ الإصدار: {meta['generated_at']}",
                    f"الفترة المشمولة: {meta['period']}",
                    f"الفئات المشمولة: {meta['categories']}",
                ):
                    c.drawRightString(width - margin, y, rs(line))
                    y -= 0.45 * cm
            else:
                c.setLineWidth(0.8)
                c.setStrokeColorRGB(0.75, 0.75, 0.75)
                c.line(margin, y, width - margin, y); y -= 0.4 * cm

            return y - 0.2 * cm

        def new_page(is_first=False):
            if not is_first:
                draw_footer()
                c.showPage()
                page_num[0] += 1
            return draw_header(is_first)

        def ensure_space(y, needed):
            if y - needed < margin + 1 * cm:
                return new_page(False)
            return y

        y = new_page(is_first=True)

        # ── Summary stat boxes ───────────────────────────────────────────
        total_donations = sum(d.amount for d in donations)
        total_expenses  = sum(e.amount for e in expenses)
        net = total_donations - total_expenses
        stats = [
            ('إجمالي التبرعات', f"{total_donations} د.ج"),
            ('إجمالي المصروفات', f"{total_expenses} د.ج"),
            ('الصافي', f"{net} د.ج"),
        ]
        gap = 0.4 * cm
        box_w = (width - 2 * margin - 2 * gap) / 3
        box_h = 1.6 * cm
        y -= box_h
        for i, (label, val) in enumerate(stats):
            bx = margin + i * (box_w + gap)
            c.setStrokeColorRGB(0.7, 0.7, 0.7)
            c.setLineWidth(0.8)
            c.roundRect(bx, y, box_w, box_h, 4, stroke=1, fill=0)
            c.setFont(arabic_font, 9)
            c.drawCentredString(bx + box_w / 2, y + box_h - 0.55 * cm, rs(label))
            c.setFont(arabic_font, 12)
            c.drawCentredString(bx + box_w / 2, y + 0.45 * cm, rs(val))
        y -= 0.7 * cm

        # ── Table drawing helper ─────────────────────────────────────────
        def draw_table(section_title, headers, col_widths, rows):
            nonlocal y
            y = ensure_space(y, 1.2 * cm)
            c.setFont(arabic_font, 11.5)
            c.drawRightString(width - margin, y, rs(section_title))
            y -= 0.5 * cm

            row_h = 0.55 * cm
            table_w = sum(col_widths)
            x_start = width - margin - table_w

            def draw_header_row():
                nonlocal y
                c.setFillColorRGB(0.078, 0.722, 0.651)  # teal, matches app branding
                c.rect(x_start, y - row_h, table_w, row_h, stroke=0, fill=1)
                c.setFillColorRGB(1, 1, 1)
                c.setFont(arabic_font, 9)
                cx = width - margin
                for h, w in zip(headers, col_widths):
                    c.drawCentredString(cx - w / 2, y - row_h + 0.16 * cm, rs(h))
                    cx -= w
                c.setFillColorRGB(0, 0, 0)
                y -= row_h

            draw_header_row()
            c.setFont(arabic_font, 8.5)
            for ridx, row in enumerate(rows):
                if y - row_h < margin + 1 * cm:
                    y = new_page(False)
                    y = ensure_space(y, row_h)
                    draw_header_row()
                if ridx % 2 == 0:
                    c.setFillColorRGB(0.95, 0.97, 0.97)
                    c.rect(x_start, y - row_h, table_w, row_h, stroke=0, fill=1)
                    c.setFillColorRGB(0, 0, 0)
                cx = width - margin
                for val, w in zip(row, col_widths):
                    c.drawCentredString(cx - w / 2, y - row_h + 0.16 * cm, rs(str(val)))
                    cx -= w
                c.setStrokeColorRGB(0.85, 0.85, 0.85)
                c.setLineWidth(0.4)
                c.line(x_start, y - row_h, x_start + table_w, y - row_h)
                y -= row_h

            c.setStrokeColorRGB(0.6, 0.6, 0.6)
            c.setLineWidth(0.8)
            c.rect(x_start, y, table_w, (len(rows) + 1) * row_h, stroke=1, fill=0)
            y -= 0.6 * cm

        donation_widths = [1 * cm, 3.4 * cm, 2.6 * cm, 2.6 * cm, 3.4 * cm, 2.6 * cm]
        draw_table(
            'سجل التبرعات',
            ['#', 'المتبرع', 'المبلغ (د.ج)', 'طريقة الدفع', 'الفئة', 'التاريخ'],
            donation_widths,
            [[i + 1, d.donor_name, d.amount, d.get_payment_method_display(), d.category.name, str(d.date)]
             for i, d in enumerate(donations)] or [['—', '—', '—', '—', '—', '—']],
        )

        expense_widths = [1 * cm, 2.8 * cm, 3.2 * cm, 5.6 * cm, 2.8 * cm]
        draw_table(
            'سجل المصروفات',
            ['#', 'المبلغ (د.ج)', 'الفئة', 'الوصف', 'التاريخ'],
            expense_widths,
            [[i + 1, e.amount, e.category.name, e.description[:40], str(e.date)]
             for i, e in enumerate(expenses)] or [['—', '—', '—', '—', '—']],
        )

        draw_footer()
        c.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="financial_report.pdf"'
        return response


class FinancialAuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = FinancialAuditLog.objects.select_related('user').all()
    serializer_class   = FinancialAuditLogSerializer
    permission_classes = [HasActivityAccess]
    activity_keywords  = FINANCE_KEYWORDS

ORG_NAME = 'جمعية الغد الأفضل'


def _find_logo_path():
    for rel in [('static', 'images', 'logo.jpg'), ('static', 'images', 'logo.png'),
                ('static', 'img', 'logo.jpg'), ('static', 'img', 'logo.png')]:
        p = os.path.join(settings.BASE_DIR, *rel)
        if os.path.exists(p):
            return p
    return None


def _get_arabic_font():
    arabic_font = 'Helvetica'
    try:
        possible_font_paths = [
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf'),
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NotoSansArabic-Regular.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        ]
        for fp in possible_font_paths:
            if os.path.exists(fp):
                pdfmetrics.registerFont(TTFont('ArabicFont', fp))
                arabic_font = 'ArabicFont'
                break
    except Exception:
        pass
    return arabic_font


def _get_reshaper():
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return lambda t: get_display(arabic_reshaper.reshape(str(t)))
    except ImportError:
        return lambda t: str(t)

ORG_NAME = 'جمعية الغد الأفضل'
ORG_ADDRESS = 'العنوان: حمام بوغرارة (ولاية تلمسان) | الهاتف: 0553497821'
# ^ update this with your organization's real address/contact info


def _draw_receipt_cell(c, x, y, w, h, item, rs, arabic_font, logo_path):
    """
    Draws a single, larger, more polished receipt inside the cell rectangle
    whose bottom-left corner is (x, y) and whose size is (w, h).
    """
    pad = 0.3 * cm
    inner_x = x + pad
    inner_y = y + pad
    inner_w = w - 2 * pad
    inner_h = h - 2 * pad

    TEAL = (0.078, 0.722, 0.651)  # matches the app's branding color used elsewhere

    # ── Outer card ────────────────────────────────────────────────────
    c.setStrokeColorRGB(*TEAL)
    c.setLineWidth(1.3)
    c.roundRect(inner_x, inner_y, inner_w, inner_h, 8, stroke=1, fill=0)

    # ── Header band (logo + org name + receipt title) ───────────────────
    header_h = inner_h * 0.17
    header_y = inner_y + inner_h - header_h

    c.saveState()
    p = c.beginPath()
    p.roundRect(inner_x, inner_y, inner_w, inner_h, 8)
    c.clipPath(p, stroke=0, fill=0)
    c.setFillColorRGB(*TEAL)
    c.rect(inner_x, header_y, inner_w, header_h, stroke=0, fill=1)
    c.restoreState()

    if logo_path:
        try:
            logo_size = min(header_h * 0.75, 1.5 * cm)
            logo = ImageReader(logo_path)
            c.drawImage(
                logo, inner_x + 0.25 * cm, header_y + (header_h - logo_size) / 2,
                width=logo_size, height=logo_size, mask='auto', preserveAspectRatio=True
            )
        except Exception:
            pass

    c.setFillColorRGB(1, 1, 1)
    c.setFont(arabic_font, min(14, h * 0.11))
    c.drawCentredString(inner_x + inner_w / 2, header_y + header_h * 0.60, rs(ORG_NAME))
    c.setFont(arabic_font, min(11, h * 0.08))
    c.drawCentredString(inner_x + inner_w / 2, header_y + header_h * 0.18, rs(item['title']))
    c.setFillColorRGB(0, 0, 0)

    # ── Body fields ───────────────────────────────────────────────────
    cy = header_y - 0.6 * cm
    right_x = inner_x + inner_w - 0.5 * cm
    label_font = min(9.5, h * 0.068)
    value_font = min(12.5, h * 0.09)

    def field(label, value):
        nonlocal cy
        c.setFont(arabic_font, label_font)
        c.setFillColorRGB(0.45, 0.45, 0.45)
        c.drawRightString(right_x, cy, rs(label))
        cy -= 0.44 * cm
        c.setFont(arabic_font, value_font)
        c.setFillColorRGB(0.1, 0.1, 0.1)
        c.drawRightString(right_x, cy - 0.1 * cm, rs(str(value)))
        cy -= 0.32 * cm
        c.setStrokeColorRGB(0.85, 0.85, 0.85)
        c.setLineWidth(0.5)
        c.line(inner_x + 0.4 * cm, cy, right_x, cy)
        cy -= 0.4 * cm

    if item.get('ref'):
        field('الرقم المرجعي', item['ref'])
    field(item['name_label'], item['name_value'])
    if item.get('method_label'):
        field(item['method_label'], item['method_value'])
    field(item['extra_label'], item['extra_value'])
    field('التاريخ', str(item['date_value']))

# ── Highlighted amount box ───────────────────────────────────────
    amount_box_h = 1.15 * cm
    cy -= 0.15 * cm
    c.setFillColorRGB(0.93, 0.98, 0.97)
    c.setStrokeColorRGB(*TEAL)
    c.setLineWidth(1)
    c.roundRect(inner_x + 0.4 * cm, cy - amount_box_h, inner_w - 0.8 * cm, amount_box_h, 5, stroke=1, fill=1)
    c.setFont(arabic_font, min(9, h * 0.06))
    c.setFillColorRGB(0.2, 0.2, 0.2)
    c.setFont(arabic_font, min(16, h * 0.115))
    c.setFillColorRGB(*TEAL)
    c.drawCentredString(inner_x + inner_w / 2, cy - amount_box_h + 0.45 * cm, rs(f"{item['amount']} دج"))
    c.setFillColorRGB(0, 0, 0)
    cy -= amount_box_h + 0.35 * cm

# ── Footer: organization address / phone line ─────────────────────
    c.setStrokeColorRGB(0.7, 0.7, 0.7)
    c.setLineWidth(0.5)
    c.line(inner_x + 0.4 * cm, cy, right_x, cy)
    cy -= 0.32 * cm
    c.setFont(arabic_font, min(6.5, h * 0.045))
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawCentredString(inner_x + inner_w / 2, cy, rs(ORG_ADDRESS))
    c.setFillColorRGB(0, 0, 0)
    cy -= 0.55 * cm

# ── Signature area (centered, directly below the address line) ────
    sig_line_w = inner_w * 0.42
    center_x = inner_x + inner_w / 2
    c.setFont(arabic_font, min(8.5, h * 0.06))
    c.setFillColorRGB(0.3, 0.3, 0.3)
    c.drawCentredString(center_x, cy + 0.10 * cm, rs('التوقيع والختم'))
    c.setStrokeColorRGB(0.5, 0.5, 0.5)
    c.setLineWidth(0.6)
    c.line(center_x - sig_line_w / 2, cy- 0.1 * cm, center_x + sig_line_w/ 2, cy - 0.1 * cm)
    c.setFillColorRGB(0, 0, 0)

def generate_receipts_pdf(items, filename='receipts.pdf'):
    """
    Lays out up to 4 receipts per A4 page in a 2x2 grid, with margins and
    dashed cutting lines between cells, so a printed sheet can be cut into
    4 individual receipts. A single receipt (items list of length 1) still
    renders on a full A4 page — occupying one real quarter of it, with the
    other three cells left as blank cut-out cells — instead of being
    stretched to fill (or shrunk to waste) an entire sheet by itself.
    """
    width, height = A4
    margin = 1 * cm
    cols, rows_per_page = 2, 2
    per_page = cols * rows_per_page

    cell_w = (width - 2 * margin) / cols
    cell_h = (height - 2 * margin) / rows_per_page

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)

    arabic_font = _get_arabic_font()
    rs = _get_reshaper()
    logo_path = _find_logo_path()

    def draw_cut_lines():
        c.setDash(3, 3)
        c.setStrokeColorRGB(0.6, 0.6, 0.6)
        c.setLineWidth(0.6)
        # vertical divider
        c.line(width / 2, margin, width / 2, height - margin)
        # horizontal divider
        c.line(margin, height / 2, width - margin, height / 2)
        c.setDash()  # reset to solid

    for idx, item in enumerate(items):
        pos_in_page = idx % per_page
        if pos_in_page == 0:
            if idx != 0:
                c.showPage()
            draw_cut_lines()
        col = pos_in_page % cols
        row = pos_in_page // cols
        x = margin + col * cell_w
        y = height - margin - (row + 1) * cell_h
        _draw_receipt_cell(c, x, y, cell_w, cell_h, item, rs, arabic_font, logo_path)

    c.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response